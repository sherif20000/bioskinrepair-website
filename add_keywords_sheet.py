from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_BG = "1A1A2E"
HEADER_FG = "FFFFFF"
ALT_ROW = "F0F4FF"
WHITE = "FFFFFF"

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Top 150 keywords for the skin barrier repair niche
# Columns: Keyword | Search Intent | Funnel Stage | Priority | Notes
keywords = [
    # --- Core / Pillar ---
    ("skin barrier repair", "Informational", "Top", "P1", "Primary niche keyword"),
    ("skin barrier", "Informational", "Top", "P1", "Broadest head term"),
    ("damaged skin barrier", "Informational", "Top", "P1", "High volume intent to learn"),
    ("how to repair skin barrier", "Informational", "Top", "P1", "Top how-to query"),
    ("skin barrier function", "Informational", "Top", "P1", "Educational intent"),
    ("what is the skin barrier", "Informational", "Top", "P1", "TOFU educational"),
    ("skin barrier damage symptoms", "Informational", "Top", "P1", "Problem-aware user"),
    ("how to fix damaged skin barrier", "Informational", "Top", "P1", "Action intent"),
    ("skin barrier recovery", "Informational", "Top", "P1", "Recovery phase"),
    ("restore skin barrier", "Informational", "Top", "P1", "Active repair intent"),

    # --- Healing & Timeline ---
    ("how long does skin barrier take to heal", "Informational", "Mid", "P1", "High commercial & info overlap"),
    ("signs skin barrier is healing", "Informational", "Mid", "P1", "Progress checker"),
    ("skin barrier repair timeline", "Informational", "Mid", "P2", ""),
    ("how to speed up skin barrier repair", "Informational", "Mid", "P1", "High intent"),
    ("why does skin feel tight", "Informational", "Top", "P2", "Symptom-led entry"),
    ("why is my skin flaky", "Informational", "Top", "P2", "Symptom-led entry"),
    ("skin barrier compromised", "Informational", "Mid", "P2", ""),
    ("how to tell if skin barrier is damaged", "Informational", "Top", "P1", "Diagnosis intent"),
    ("skin barrier healing signs week by week", "Informational", "Mid", "P2", ""),
    ("skin barrier still damaged after weeks", "Informational", "Mid", "P2", "Frustrated user"),

    # --- Ingredients (humectants, ceramides, occlusives, emollients) ---
    ("ceramides for skin", "Informational", "Mid", "P1", "Core ingredient term"),
    ("ceramide moisturizer", "Commercial", "Mid", "P1", "Buying intent"),
    ("best ceramide moisturizer", "Commercial", "Bottom", "P1", "High purchase intent"),
    ("ceramide cream for dry skin", "Commercial", "Bottom", "P1", ""),
    ("what do ceramides do for skin", "Informational", "Top", "P2", ""),
    ("humectant skincare", "Informational", "Mid", "P2", ""),
    ("what is a humectant", "Informational", "Top", "P2", ""),
    ("best humectant for dry skin", "Commercial", "Bottom", "P2", ""),
    ("humectant vs emollient vs occlusive", "Informational", "Mid", "P1", "Comparison intent"),
    ("occlusive skincare", "Informational", "Mid", "P2", ""),
    ("what is an occlusive", "Informational", "Top", "P2", ""),
    ("best occlusive moisturizer", "Commercial", "Bottom", "P2", ""),
    ("occlusive for skin barrier", "Informational", "Mid", "P2", ""),
    ("emollient skin", "Informational", "Mid", "P2", ""),
    ("what is an emollient in skincare", "Informational", "Top", "P2", ""),
    ("emollient vs moisturizer", "Informational", "Mid", "P2", ""),
    ("hyaluronic acid for skin barrier", "Informational", "Mid", "P1", ""),
    ("hyaluronic acid dry skin", "Informational", "Mid", "P1", ""),
    ("is hyaluronic acid good for damaged skin", "Informational", "Mid", "P1", ""),
    ("niacinamide skin barrier", "Informational", "Mid", "P1", ""),
    ("niacinamide for sensitive skin", "Informational", "Mid", "P1", ""),
    ("glycerin for skin", "Informational", "Mid", "P2", ""),
    ("glycerin vs hyaluronic acid", "Informational", "Mid", "P2", ""),
    ("panthenol skincare", "Informational", "Mid", "P2", ""),
    ("vitamin B5 skin", "Informational", "Mid", "P2", ""),
    ("TEWL skin barrier", "Informational", "Mid", "P2", "Niche technical term"),
    ("transepidermal water loss", "Informational", "Mid", "P2", ""),

    # --- Skincare Routines ---
    ("skincare routine for damaged skin barrier", "Informational", "Mid", "P1", "High purchase proximity"),
    ("minimalist skincare routine sensitive skin", "Informational", "Mid", "P1", ""),
    ("simple skincare routine barrier repair", "Informational", "Mid", "P1", ""),
    ("what not to use on damaged skin barrier", "Informational", "Mid", "P1", "Negative intent — high click"),
    ("skincare routine skin barrier recovery", "Informational", "Mid", "P1", ""),
    ("how many products to use with damaged barrier", "Informational", "Mid", "P2", ""),
    ("can you use retinol with damaged skin barrier", "Informational", "Mid", "P1", ""),
    ("how to layer skincare products", "Informational", "Top", "P2", ""),

    # --- Exfoliation & Over-Exfoliation ---
    ("over exfoliation", "Informational", "Mid", "P1", "High volume problem query"),
    ("signs of over exfoliated skin", "Informational", "Mid", "P1", ""),
    ("how to recover from over exfoliation", "Informational", "Mid", "P1", ""),
    ("over exfoliation recovery routine", "Informational", "Mid", "P1", ""),
    ("AHA vs BHA sensitive skin", "Informational", "Mid", "P1", ""),
    ("chemical exfoliant damaged barrier", "Informational", "Mid", "P2", ""),
    ("how often to exfoliate sensitive skin", "Informational", "Mid", "P2", ""),

    # --- Retinol ---
    ("retinol skin barrier damage", "Informational", "Mid", "P1", ""),
    ("does retinol damage skin barrier", "Informational", "Mid", "P1", ""),
    ("retinol for beginners sensitive skin", "Informational", "Mid", "P1", ""),
    ("how to use retinol without irritation", "Informational", "Mid", "P1", ""),
    ("retinol purging vs barrier damage", "Informational", "Mid", "P1", ""),
    ("when to stop retinol skin barrier", "Informational", "Mid", "P2", ""),

    # --- Sunscreen ---
    ("mineral vs chemical sunscreen", "Informational", "Mid", "P1", ""),
    ("best sunscreen for sensitive skin", "Commercial", "Bottom", "P1", ""),
    ("mineral sunscreen for damaged skin barrier", "Commercial", "Bottom", "P1", ""),
    ("does sunscreen irritate sensitive skin", "Informational", "Mid", "P1", ""),
    ("zinc oxide sunscreen sensitive skin", "Commercial", "Bottom", "P2", ""),
    ("sunscreen for skin barrier repair", "Commercial", "Bottom", "P2", ""),
    ("EltaMD UV Clear review", "Navigational", "Bottom", "P2", ""),

    # --- Condition: Eczema ---
    ("eczema skin barrier", "Informational", "Mid", "P1", ""),
    ("eczema barrier repair", "Informational", "Mid", "P1", ""),
    ("best moisturizer for eczema", "Commercial", "Bottom", "P1", ""),
    ("ceramides for eczema", "Informational", "Mid", "P1", ""),
    ("eczema skincare routine", "Informational", "Mid", "P1", ""),
    ("what causes eczema flares", "Informational", "Top", "P2", ""),

    # --- Condition: Rosacea ---
    ("rosacea skin barrier", "Informational", "Mid", "P1", ""),
    ("best moisturizer for rosacea", "Commercial", "Bottom", "P1", ""),
    ("rosacea skincare routine", "Informational", "Mid", "P1", ""),
    ("does rosacea damage skin barrier", "Informational", "Mid", "P2", ""),
    ("rosacea barrier repair routine", "Informational", "Mid", "P2", ""),

    # --- Condition: Perioral Dermatitis ---
    ("perioral dermatitis", "Informational", "Mid", "P1", ""),
    ("perioral dermatitis causes", "Informational", "Mid", "P1", ""),
    ("perioral dermatitis treatment", "Commercial", "Bottom", "P1", ""),
    ("how to clear perioral dermatitis", "Informational", "Mid", "P1", ""),
    ("perioral dermatitis skincare routine", "Informational", "Mid", "P2", ""),
    ("perioral dermatitis vs rosacea", "Informational", "Mid", "P2", ""),

    # --- Condition: Fungal Acne ---
    ("fungal acne", "Informational", "Mid", "P1", ""),
    ("malassezia folliculitis", "Informational", "Mid", "P2", ""),
    ("fungal acne vs regular acne", "Informational", "Mid", "P1", ""),
    ("fungal acne safe moisturizer", "Commercial", "Bottom", "P1", ""),
    ("how to treat fungal acne", "Informational", "Mid", "P1", ""),
    ("fungal acne skincare routine", "Informational", "Mid", "P2", ""),

    # --- Sensitive Skin ---
    ("sensitive skin vs damaged skin barrier", "Informational", "Mid", "P1", ""),
    ("how to know if skin is sensitive", "Informational", "Top", "P2", ""),
    ("best moisturizer for sensitive skin", "Commercial", "Bottom", "P1", ""),
    ("sensitive skin routine", "Informational", "Mid", "P1", ""),
    ("non-irritating skincare sensitive skin", "Commercial", "Bottom", "P2", ""),

    # --- Acne & Barrier ---
    ("skin barrier repair acne prone skin", "Informational", "Mid", "P1", ""),
    ("does acne damage skin barrier", "Informational", "Mid", "P1", ""),
    ("non-comedogenic barrier repair moisturizer", "Commercial", "Bottom", "P1", ""),
    ("barrier repair routine for acne", "Informational", "Mid", "P2", ""),
    ("best moisturizer for acne and dry skin", "Commercial", "Bottom", "P2", ""),

    # --- Product Comparisons ---
    ("CeraVe vs Cetaphil", "Commercial", "Bottom", "P1", "Very high volume"),
    ("CeraVe vs Cetaphil for eczema", "Commercial", "Bottom", "P1", ""),
    ("CeraVe vs Cetaphil sensitive skin", "Commercial", "Bottom", "P1", ""),
    ("niacinamide vs vitamin C", "Informational", "Mid", "P1", ""),
    ("can you mix niacinamide and vitamin C", "Informational", "Mid", "P1", ""),
    ("hyaluronic acid vs glycerin moisturizer", "Informational", "Mid", "P2", ""),
    ("Cicaplast B5 vs CeraVe", "Commercial", "Bottom", "P2", ""),
    ("Paula's Choice vs CeraVe", "Commercial", "Bottom", "P2", ""),

    # --- Product Reviews ---
    ("CeraVe Moisturizing Cream review", "Navigational", "Bottom", "P1", ""),
    ("La Roche-Posay Cicaplast review", "Navigational", "Bottom", "P1", ""),
    ("Paula's Choice Barrier Repair review", "Navigational", "Bottom", "P1", ""),
    ("The Ordinary NMF review", "Navigational", "Bottom", "P2", ""),
    ("Vanicream Gentle Cleanser review", "Navigational", "Bottom", "P2", ""),
    ("best barrier repair moisturizer 2026", "Commercial", "Bottom", "P1", ""),

    # --- Cleansers ---
    ("best cleanser for damaged skin barrier", "Commercial", "Bottom", "P1", ""),
    ("gentle cleanser sensitive skin", "Commercial", "Bottom", "P1", ""),
    ("sulfate-free cleanser barrier repair", "Commercial", "Bottom", "P1", ""),
    ("non-stripping face wash", "Commercial", "Bottom", "P2", ""),
    ("best face wash dry damaged skin", "Commercial", "Bottom", "P2", ""),

    # --- Broader skin health (supporting cluster) ---
    ("skin microbiome", "Informational", "Top", "P3", "Adjacent cluster"),
    ("skin pH balance", "Informational", "Top", "P3", "Supporting concept"),
    ("acid mantle skin", "Informational", "Mid", "P3", "Supporting concept"),
    ("what is the acid mantle", "Informational", "Top", "P3", ""),
    ("stratum corneum", "Informational", "Mid", "P3", "Technical term"),
    ("corneum repair", "Informational", "Mid", "P3", ""),
    ("skin barrier lipids", "Informational", "Mid", "P3", ""),
    ("epidermal barrier function", "Informational", "Mid", "P3", "Clinical term"),
    ("skin barrier genes", "Informational", "Top", "P3", "Low competition niche"),
    ("filaggrin skin barrier", "Informational", "Mid", "P3", "Genetic/clinical"),

    # --- Lifestyle & triggers ---
    ("hot water skin barrier damage", "Informational", "Mid", "P2", ""),
    ("does washing face too much damage skin", "Informational", "Mid", "P2", ""),
    ("stress skin barrier", "Informational", "Mid", "P2", ""),
    ("alcohol in skincare skin barrier", "Informational", "Mid", "P2", ""),
    ("fragrance free skincare sensitive skin", "Commercial", "Bottom", "P2", ""),
    ("does diet affect skin barrier", "Informational", "Top", "P2", ""),
    ("omega 3 for skin barrier", "Informational", "Mid", "P2", ""),
    ("sleep and skin barrier repair", "Informational", "Top", "P3", ""),

    # --- Additional P1/P2 high-value ---
    ("skin barrier 101", "Informational", "Top", "P1", "Exact-match pillar title"),
    ("barrier cream for face", "Commercial", "Bottom", "P1", ""),
    ("best moisturizer for dry damaged skin", "Commercial", "Bottom", "P1", ""),
    ("skin barrier ingredients to avoid", "Informational", "Mid", "P1", "Negative-intent high CTR"),
    ("repair skin barrier overnight", "Informational", "Mid", "P1", "Urgency intent"),
]

FILE = "/Users/sherifelkady/Claude Code Server's Projects/bioskinrepair-website/bioskinrepair-content-master.xlsx"

wb = load_workbook(FILE)

# Remove existing sheet if re-running
if "Top 150 Keywords" in wb.sheetnames:
    del wb["Top 150 Keywords"]

ws3 = wb.create_sheet("Top 150 Keywords")

headers = ["#", "Keyword", "Search Intent", "Funnel Stage", "Priority", "Notes"]
col_widths = [5, 52, 18, 14, 10, 40]

# Header row
for col, h in enumerate(headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = Font(name="Arial", bold=True, color=HEADER_FG, size=11)
    cell.fill = PatternFill("solid", start_color=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws3.row_dimensions[1].height = 30

# Intent color mapping
intent_colors = {
    "Informational": "1565C0",
    "Commercial":    "2E7D32",
    "Navigational":  "6A1B9A",
}
funnel_colors = {
    "Top":    "E65100",
    "Mid":    "1565C0",
    "Bottom": "2E7D32",
}
priority_colors = {
    "P1": "D32F2F",
    "P2": "1565C0",
    "P3": "757575",
}

for row_idx, (kw, intent, funnel, priority, notes) in enumerate(keywords, 2):
    bg = WHITE if row_idx % 2 == 0 else ALT_ROW

    def cell(col, val, bold=False, color="333333", halign="left"):
        c = ws3.cell(row=row_idx, column=col, value=val)
        c.font = Font(name="Arial", size=10, bold=bold, color=color)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
        c.border = border
        return c

    cell(1, row_idx - 1, halign="center", color="999999")
    cell(2, kw, bold=True, color="111111")
    cell(3, intent, bold=True, color=intent_colors.get(intent, "333333"), halign="center")
    cell(4, funnel, color=funnel_colors.get(funnel, "333333"), halign="center")
    cell(5, priority, bold=True, color=priority_colors.get(priority, "333333"), halign="center")
    cell(6, notes, color="666666")

    ws3.row_dimensions[row_idx].height = 18

# Column widths
for i, w in enumerate(col_widths, 1):
    from openpyxl.utils import get_column_letter
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.freeze_panes = "A2"
ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# Move sheet to position 2 (after Content Inventory)
wb.move_sheet("Top 150 Keywords", offset=-(len(wb.sheetnames) - 2))

wb.save(FILE)
print(f"Done — {len(keywords)} keywords added to 'Top 150 Keywords' sheet")
